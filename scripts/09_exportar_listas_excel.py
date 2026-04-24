"""
Exporta listas de precios formateadas a Excel con estilos profesionales.
Genera tres hojas: Lista Publica, Lista Distribuidor, Insumos.
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from sqlalchemy import create_engine, text
from dotenv import load_dotenv
from urllib.parse import quote_plus
import os, datetime

load_dotenv()
password = quote_plus(os.getenv('DB_PASSWORD'))
engine = create_engine(
    f"postgresql://{os.getenv('DB_USER')}:{password}@"
    f"{os.getenv('DB_HOST')}:{os.getenv('DB_PORT')}/{os.getenv('DB_NAME')}"
)

# ── Estilos ───────────────────────────────────────────────────────────────────
AZUL_OSCURO  = "1F497D"
AZUL_CLARO   = "DEEAF1"
GRIS         = "F2F2F2"
VERDE_OSCURO = "375623"
VERDE_CLARO  = "E2EFDA"
NARANJA      = "C55A11"
NARANJA_CLAR = "FCE4D6"
BLANCO       = "FFFFFF"
NEGRO        = "000000"

def header_fill(color):  return PatternFill("solid", fgColor=color)
def row_fill(color):     return PatternFill("solid", fgColor=color)
def bold_font(color=NEGRO, size=11): return Font(bold=True, color=color, size=size)
def normal_font(size=10): return Font(size=size)

thin = Side(style="thin", color="CCCCCC")
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

def auto_width(ws, min_w=8, max_w=50):
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = \
            min(max_w, max(min_w, max_len + 2))

def write_header_row(ws, row_num, labels, fill_color, font_color=BLANCO):
    for col_num, label in enumerate(labels, 1):
        cell = ws.cell(row=row_num, column=col_num, value=label)
        cell.font      = bold_font(font_color, 10)
        cell.fill      = header_fill(fill_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border

def write_title(ws, title, subtitle, fill_color):
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value     = title
    t.font      = Font(bold=True, size=16, color=BLANCO)
    t.fill      = header_fill(fill_color)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    s = ws["A2"]
    s.value     = subtitle
    s.font      = Font(size=10, color="666666")
    s.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

def format_number(ws, row, col, value, fmt="$#,##0.00"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.number_format = fmt
    cell.alignment = Alignment(horizontal="right")
    cell.border = thin_border
    return cell

def format_text(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = Alignment(vertical="center")
    cell.border = thin_border
    return cell

# ─────────────────────────────────────────────────────────────────────────────
# Cargar datos
# ─────────────────────────────────────────────────────────────────────────────
print("Cargando datos...")
with engine.connect() as conn:
    df_pub  = pd.read_sql(text("SELECT * FROM v_lista_precios_publica ORDER BY categoria, nombre"), conn)
    df_dist = pd.read_sql(text("SELECT * FROM v_lista_precios_distribuidor ORDER BY categoria, nombre"), conn)
    df_ins  = pd.read_sql(text("""
        SELECT codigo, descripcion, rubro, proveedor,
               precio_ars, precio_ars_iva, precio_usd,
               cotizacion_dolar, stock
        FROM insumos
        WHERE precio_ars IS NOT NULL
        ORDER BY rubro, descripcion
    """), conn)

wb = Workbook()
fecha_str = datetime.date.today().strftime("%d/%m/%Y")

# ═════════════════════════════════════════════════════════════════════════════
# HOJA 1 — Lista Pública
# ═════════════════════════════════════════════════════════════════════════════
ws_pub = wb.active
ws_pub.title = "Lista Publica"

write_title(ws_pub, "ELEKTRA S.R.L. — Lista de Precios al Público",
            f"Actualizada al {fecha_str}  |  Precios en ARS con IVA incluido", AZUL_OSCURO)

headers_pub = ["#", "Categoría", "Producto", "Sin IVA ($)", "IVA (%)", "Con IVA ($)", "USD", "Dto. Dist. (%)"]
write_header_row(ws_pub, 3, headers_pub, AZUL_OSCURO)
ws_pub.row_dimensions[3].height = 22

categoria_actual = None
fila = 4
for i, row in df_pub.iterrows():
    # Fila de separación por categoría
    if row["categoria"] != categoria_actual:
        categoria_actual = row["categoria"]
        ws_pub.merge_cells(f"A{fila}:H{fila}")
        cell = ws_pub.cell(row=fila, column=1, value=f"  {categoria_actual}")
        cell.font = bold_font(BLANCO, 10)
        cell.fill = header_fill("2E74B5")
        cell.alignment = Alignment(vertical="center")
        ws_pub.row_dimensions[fila].height = 18
        fila += 1

    fill = row_fill(AZUL_CLARO) if fila % 2 == 0 else row_fill(BLANCO)

    format_text(ws_pub,  fila, 1, i + 1).fill = fill
    format_text(ws_pub,  fila, 2, row["categoria"]).fill = fill
    format_text(ws_pub,  fila, 3, row["nombre"]).fill = fill
    format_number(ws_pub, fila, 4, row["precio_sin_iva"]).fill = fill
    c = ws_pub.cell(row=fila, column=5, value=int(row["iva_pct"]) if row["iva_pct"] else 21)
    c.number_format = '0"%"'
    c.alignment = Alignment(horizontal="center")
    c.border = thin_border
    c.fill = fill
    format_number(ws_pub, fila, 6, row["precio_con_iva"]).fill = fill
    format_number(ws_pub, fila, 7, row["precio_usd"], "$#,##0.00").fill = fill
    c2 = ws_pub.cell(row=fila, column=8, value=row["descuento_dist_pct"])
    c2.number_format = '0"%"'
    c2.alignment = Alignment(horizontal="center")
    c2.border = thin_border
    c2.fill = fill
    fila += 1

auto_width(ws_pub)
ws_pub.column_dimensions["C"].width = 45
ws_pub.freeze_panes = "A4"

# ═════════════════════════════════════════════════════════════════════════════
# HOJA 2 — Lista Distribuidor
# ═════════════════════════════════════════════════════════════════════════════
ws_dist = wb.create_sheet("Lista Distribuidor")

write_title(ws_dist, "ELEKTRA S.R.L. — Lista de Precios Distribuidor",
            f"Actualizada al {fecha_str}  |  USO INTERNO — NO DIFUNDIR", VERDE_OSCURO)

headers_dist = ["#", "Categoría", "Producto", "Sin IVA ($)", "Con IVA ($)", "USD ref.", "Dto. (%)"]
write_header_row(ws_dist, 3, headers_dist, VERDE_OSCURO)
ws_dist.row_dimensions[3].height = 22

categoria_actual = None
fila = 4
for i, row in df_dist.iterrows():
    if row["categoria"] != categoria_actual:
        categoria_actual = row["categoria"]
        ws_dist.merge_cells(f"A{fila}:G{fila}")
        cell = ws_dist.cell(row=fila, column=1, value=f"  {categoria_actual}")
        cell.font = bold_font(BLANCO, 10)
        cell.fill = header_fill("538135")
        cell.alignment = Alignment(vertical="center")
        ws_dist.row_dimensions[fila].height = 18
        fila += 1

    fill = row_fill(VERDE_CLARO) if fila % 2 == 0 else row_fill(BLANCO)
    format_text(ws_dist,   fila, 1, i + 1).fill = fill
    format_text(ws_dist,   fila, 2, row["categoria"]).fill = fill
    format_text(ws_dist,   fila, 3, row["nombre"]).fill = fill
    format_number(ws_dist, fila, 4, row["precio_dist_sin_iva"]).fill = fill
    format_number(ws_dist, fila, 5, row["precio_dist_con_iva"]).fill = fill
    format_number(ws_dist, fila, 6, row["precio_referencia_usd"]).fill = fill
    c = ws_dist.cell(row=fila, column=7, value=row["descuento_pct"])
    c.number_format = '0"%"'
    c.alignment = Alignment(horizontal="center")
    c.border = thin_border
    c.fill = fill
    fila += 1

auto_width(ws_dist)
ws_dist.column_dimensions["C"].width = 45
ws_dist.freeze_panes = "A4"

# ═════════════════════════════════════════════════════════════════════════════
# HOJA 3 — Insumos
# ═════════════════════════════════════════════════════════════════════════════
ws_ins = wb.create_sheet("Insumos")

write_title(ws_ins, "ELEKTRA S.R.L. — Lista de Insumos y Materiales",
            f"Actualizada al {fecha_str}  |  Precios en ARS", NARANJA)

headers_ins = ["Código", "Rubro", "Descripción", "Proveedor", "Precio ARS", "c/IVA ($)", "USD", "Cotiz. $", "Stock"]
write_header_row(ws_ins, 3, headers_ins, NARANJA)
ws_ins.row_dimensions[3].height = 22

rubro_actual = None
fila = 4
for _, row in df_ins.iterrows():
    if row["rubro"] != rubro_actual:
        rubro_actual = row["rubro"]
        ws_ins.merge_cells(f"A{fila}:I{fila}")
        cell = ws_ins.cell(row=fila, column=1, value=f"  {rubro_actual}")
        cell.font = bold_font(BLANCO, 10)
        cell.fill = header_fill("ED7D31")
        cell.alignment = Alignment(vertical="center")
        ws_ins.row_dimensions[fila].height = 18
        fila += 1

    fill = row_fill(NARANJA_CLAR) if fila % 2 == 0 else row_fill(BLANCO)
    format_text(ws_ins,   fila, 1, row["codigo"]).fill = fill
    format_text(ws_ins,   fila, 2, row["rubro"]).fill = fill
    format_text(ws_ins,   fila, 3, row["descripcion"]).fill = fill
    format_text(ws_ins,   fila, 4, row["proveedor"] or "").fill = fill
    format_number(ws_ins, fila, 5, row["precio_ars"]).fill = fill
    format_number(ws_ins, fila, 6, row["precio_ars_iva"]).fill = fill
    usd = row["precio_usd"] if pd.notna(row["precio_usd"]) else None
    format_number(ws_ins, fila, 7, usd).fill = fill
    format_number(ws_ins, fila, 8, row["cotizacion_dolar"], "#,##0").fill = fill
    stk = row["stock"] if pd.notna(row["stock"]) else None
    format_number(ws_ins, fila, 9, stk, "#,##0.##").fill = fill
    fila += 1

auto_width(ws_ins)
ws_ins.column_dimensions["C"].width = 50
ws_ins.column_dimensions["D"].width = 25
ws_ins.freeze_panes = "A4"

# ── Guardar ───────────────────────────────────────────────────────────────────
fecha_archivo = datetime.date.today().strftime("%Y%m%d")
output = f"../data/ELEKTRA_Listas_Precios_{fecha_archivo}.xlsx"
wb.save(output)
print(f"Excel guardado: {output}")
print(f"  Hoja 'Lista Publica':      {len(df_pub):,} productos")
print(f"  Hoja 'Lista Distribuidor': {len(df_dist):,} productos")
print(f"  Hoja 'Insumos':            {len(df_ins):,} materiales")
